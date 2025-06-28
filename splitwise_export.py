import os
from splitwise import Splitwise
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime

# Load environment variables
load_dotenv()

CONSUMER_KEY = os.getenv("SPLITWISE_CONSUMER_KEY")
CONSUMER_SECRET = os.getenv("SPLITWISE_CONSUMER_SECRET")
ACCESS_TOKEN = os.getenv("SPLITWISE_ACCESS_TOKEN")
ACCESS_TOKEN_SECRET = os.getenv("SPLITWISE_ACCESS_TOKEN_SECRET")
GROUP_ID = os.getenv("GROUP_ID")

if not all([CONSUMER_KEY, CONSUMER_SECRET, ACCESS_TOKEN, ACCESS_TOKEN_SECRET, GROUP_ID]):
    raise Exception("One or more environment variables are missing. Please check your .env file.")

# Authenticate with Splitwise
sObj = Splitwise(CONSUMER_KEY, CONSUMER_SECRET, ACCESS_TOKEN, ACCESS_TOKEN_SECRET)

def main():
    # Fetch all expenses for the group
    expenses = []
    offset = 0
    limit = 50  # Splitwise API default limit
    while True:
        batch = sObj.getExpenses(offset=offset, limit=limit, group_id=int(GROUP_ID))
        if not batch:
            break
        expenses.extend(batch)
        if len(batch) < limit:
            break
        offset += limit
    print(f"Fetched {len(expenses)} expenses for group {GROUP_ID}.")

    # Group expenses by month
    expenses_by_month = defaultdict(list)
    for exp in expenses:
        date_str = exp.getDate()  # e.g., '2023-06-15T12:34:56Z'
        if date_str:
            date_obj = datetime.strptime(date_str[:10], "%Y-%m-%d")
            month_key = date_obj.strftime("%Y-%m")
        else:
            month_key = "Unknown"
        expenses_by_month[month_key].append(exp)

    # Create Excel workbook
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    for month, exps in sorted(expenses_by_month.items()):
        ws = wb.create_sheet(title=month)
        ws.append(["Date", "Description", "Cost", "Category"])
        for exp in exps:
            date = exp.getDate()[:10] if exp.getDate() else ""
            desc = exp.getDescription() or ""
            cost = exp.getCost() or ""
            cat = exp.getCategory().getName() if exp.getCategory() else ""
            ws.append([date, desc, cost, cat])
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

    # Save workbook
    output_file = f"splitwise_group_{GROUP_ID}_expenses.xlsx"
    wb.save(output_file)
    print(f"Exported expenses to {output_file}")

if __name__ == "__main__":
    main() 